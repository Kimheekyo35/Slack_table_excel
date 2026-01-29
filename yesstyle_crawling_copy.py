import os
import time
import re
import random
from pathlib import Path
from typing import Optional, Iterable, Tuple, List
from io import BytesIO
import pandas as pd
import psycopg2
from psycopg2 import Error
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium_utils import build_chrome_options, create_chrome_driver, cleanup_driver
from slack_sdk import WebClient
# --------------------------------------------------------------------------- #
# 환경 설정
# --------------------------------------------------------------------------- #
load_dotenv(override=True)

DB_HOST = os.getenv("PG_HOST", "141.164.49.115")
DB_PORT = os.getenv("PG_PORT", "5432")
DB_DATABASE = os.getenv("PG_DATABASE", "benow_db")
DB_USER = os.getenv("PG_USER", "wemarketing_user")
DB_PASSWORD = os.getenv("PG_PASSWORD", "ehgus1500")

SEOUL_TZ = ZoneInfo("Asia/Seoul")

# 컬럼 이름 상수
COL_RANK = "순위"
COL_BRAND = "브랜드명"
COL_PRODUCT_NAME = "제품명"
COL_PRICE = "가격"
COL_DATETIME_TEXT = "날짜와시간"
COL_CHANNEL = "채널"
COL_CATEGORY = "카테고리"
COL_COLLECTED_AT = "수집일시"

COL_PREVIOUS_RANK = "전일 순위"
COL_RANK_DELTA = "전일 변동"
COL_STATUS = "전일대비 증감"

COL_PREVIOUS_RANK_WEEK = "전주 순위"
COL_WEEK_DELTA = "전주 변동"
COL_WEEK_STATUS = "전주대비 증감"

# 크롤링 대상 URL
BESTSELLER_URL = "https://www.yesstyle.com/en/beauty-beauty/list.html/bcc.15478_bpt.46?sb=136"
SKINCARE_URL = "https://www.yesstyle.com/en/beauty-skin-care/list.html/bcc.15544_bpt.46"
NUMBUZIN_URL = "https://www.yesstyle.com/en/numbuzin/list.html/bpt.299_bid.326359"
CRAWL_LIMIT = 100

REPORT_LIMIT = 50
COL_BRAND_KEY = "_brand_key"
COL_PRODUCT_KEY = "_product_key"

# Slack API TOKEN
SLACK_BOT_TOKEN=os.getenv("SLACK_BOT_TOKEN")
# --------------------------------------------------------------------------- #
# 유틸 함수
# --------------------------------------------------------------------------- #
def ensure_data_directory() -> Path:
    base = Path(__file__).resolve().parent
    data_dir = base / "data"
    data_dir.mkdir(exist_ok=True)
    return data_dir


def fill_missing(items: List[Optional[str]], target: int) -> List[Optional[str]]:
    if len(items) < target:
        items.extend([None] * (target - len(items)))
    return items

def _extract_primary_price(text: str) -> str:
    if not text:
        return "N/A"
    match = re.search(r"([₩￦$€£¥])\s?\d[\d,]*(?:\.\d+)?", text)
    if match:
        return match.group(0).strip()
    num = re.search(r"\d[\d,]*(?:\.\d+)?", text)
    if num:
        prefix = text[:num.start()].strip()
        symbol = ""
        if prefix and prefix[-1] in "₩￦$€£¥":
            symbol = prefix[-1]
        return f"{symbol} {num.group(0)}".strip()
    return text.strip()


VOLUME_PREFIX_PATTERN = re.compile(r"^\(\d+(?:ML|EA|G|PATCHES|PCS|PADS)\)\s*", re.IGNORECASE)


def _normalize_product_name(name: str) -> str:
    if name is None:
        return ""
    text = str(name).strip()
    text = VOLUME_PREFIX_PATTERN.sub("", text)
    text = re.sub(r"\s+", " ", text)
    return text


def _normalize_match_key(value: Optional[str]) -> str:
    if value is None:
        return ""
    text = str(value)
    return re.sub(r"\s+", "", text)


def _add_match_keys_inplace(df: pd.DataFrame) -> pd.DataFrame:
    if df is None:
        return df
    if COL_BRAND in df.columns:
        df[COL_BRAND_KEY] = df[COL_BRAND].apply(_normalize_match_key)
    else:
        df[COL_BRAND_KEY] = ""
    if COL_PRODUCT_NAME in df.columns:
        df[COL_PRODUCT_KEY] = df[COL_PRODUCT_NAME].apply(_normalize_match_key)
    else:
        df[COL_PRODUCT_KEY] = ""
    return df


def _deduplicate_by_match_keys(df: Optional[pd.DataFrame], limit: Optional[int] = None) -> Optional[pd.DataFrame]:
    if df is None:
        return None
    work = df.copy()
    _add_match_keys_inplace(work)
    work = work.drop_duplicates(subset=[COL_BRAND_KEY, COL_PRODUCT_KEY])
    if limit is not None:
        work = work.head(limit)
    work.reset_index(drop=True, inplace=True)
    return work


def parse_list_grid_products(products: Iterable, start_rank: int = 1):
    names, prices, ranks, brands = [], [], [], []
    rank = start_rank
    for item in products:
        try:
            raw_name = item.find_element(By.CSS_SELECTOR, "[class*='itemTitle']").text
        except Exception:
            raw_name = "N/A"

        normalized_name = _normalize_product_name(raw_name)
        brand = "N/A"
        product = normalized_name
        if "-" in normalized_name:
            brand_part, product_part = normalized_name.split("-", 1)
            brand_candidate = brand_part.strip()
            product = _normalize_product_name(product_part)
            brand = brand_candidate if brand_candidate else "N/A"

        try:
            raw_price = item.find_element(By.CSS_SELECTOR, "[class*='itemPrice']").text
        except Exception:
            raw_price = "N/A"

        names.append(product)
        prices.append(_extract_primary_price(raw_price))
        ranks.append(rank)
        brands.append(brand)
        rank += 1
    return names, prices, ranks, brands


def yesstyle_scroll_crawling(
    driver: webdriver.Chrome,
    url: str,
    target_count: int = 100,
    product_selector: str = "a[class*='itemContainer']",
) -> Tuple[List[Optional[str]], List[Optional[str]], List[Optional[int]], List[Optional[str]]]:
    driver.get(url)
    time.sleep(random.uniform(2.5, 3.5))

    gathered_names: List[Optional[str]] = []
    gathered_prices: List[Optional[str]] = []
    gathered_ranks: List[Optional[int]] = []
    gathered_brands: List[Optional[str]] = []

    page_index = 0
    while len(gathered_names) < target_count:
        page_index += 1
        try:
            WebDriverWait(driver, 15).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, product_selector))
            )
        except Exception:
            pass

        products = driver.find_elements(By.CSS_SELECTOR, product_selector)
        if not products:
            print(f"[WARN] No products found with selector '{product_selector}' at {url}.")
            break

        remaining = target_count - len(gathered_names)
        batch = products[:remaining]
        names, prices, ranks, brands = parse_list_grid_products(
            batch, start_rank=len(gathered_names) + 1
        )
        gathered_names.extend(names)
        gathered_prices.extend(prices)
        gathered_ranks.extend(ranks)
        gathered_brands.extend(brands)

        if len(gathered_names) >= target_count:
            break

        first_old = products[0]
        if not _find_and_click_next(driver):
            break
        try:
            WebDriverWait(driver, 10).until(EC.staleness_of(first_old))
        except Exception:
            pass
        time.sleep(random.uniform(1.0, 2.0))

    fill_missing(gathered_names, target_count)
    fill_missing(gathered_prices, target_count)
    fill_missing(gathered_ranks, target_count)
    fill_missing(gathered_brands, target_count)

    return gathered_names, gathered_prices, gathered_ranks, gathered_brands


def _find_and_click_next(driver: webdriver.Chrome) -> bool:
    selectors = [
        "a[class*='nextPage']",
        "a[class*='simpleDirectionButton']",
        "button[class*='productListingMain_blackButton__']",
    ]
    for selector in selectors:
        try:
            element = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, selector))
            )
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
            time.sleep(0.2)
            element.click()
            return True
        except Exception:
            continue
    return False


# --------------------------------------------------------------------------- #
# DB 관련 함수
# --------------------------------------------------------------------------- #
def insert_into_postgresql(rows: List[Tuple]):
    if not rows:
        return
    connection = None
    try:
        connection = psycopg2.connect(
            host=DB_HOST,
            port=DB_PORT,
            database=DB_DATABASE,
            user=DB_USER,
            password=DB_PASSWORD,
        )
        cursor = connection.cursor()
        query = """
            INSERT INTO benow.yesstyle_table
                ("순위", "브랜드명", "제품명", "가격", "날짜와시간", "채널",
                 "카테고리", "정가", "판매량", "판매샵", "오특유무", created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        cursor.executemany(query, rows)
        connection.commit()
        cursor.close()
        print(f"Inserted {len(rows)} rows into yesstyle_table.")
    except Error as exc:
        if connection:
            connection.rollback()
        print(f"PostgreSQL insert error: {exc}")
    finally:
        if connection:
            connection.close()


def fetch_previous_snapshot_from_db(
    category: str,
    target_time: datetime,
    limit: Optional[int] = None,
) -> Tuple[Optional[pd.DataFrame], Optional[datetime], Optional[str]]:
    connection = None
    try:
        connection = psycopg2.connect(
            host=DB_HOST,
            port=DB_PORT,
            database=DB_DATABASE,
            user=DB_USER,
            password=DB_PASSWORD,
        )
        cursor = connection.cursor()

        target_time = target_time.replace(tzinfo=None)
        target_date = target_time.date()
        desired_time = target_time.replace(hour=8, minute=0, second=0, microsecond=0)

        cursor.execute(
            """
            SELECT created_at
            FROM benow.yesstyle_table
            WHERE "채널" = %s
              AND "카테고리" = %s
              AND created_at::date = %s
            ORDER BY created_at
            """,
            ("YesStyle", category, target_date),
        )
        day_rows = [row[0] for row in cursor.fetchall()]
        if day_rows:
            latest_created = min(day_rows, key=lambda ts: abs(ts - desired_time))
        else:
            cursor.execute(
                """
                SELECT MAX(created_at)
                FROM benow.yesstyle_table
                WHERE "채널" = %s
                  AND "카테고리" = %s
                  AND created_at < %s
                """,
                ("YesStyle", category, target_time),
            )
            latest_created = cursor.fetchone()[0]
            if latest_created is None:
                return None, None, None

        base_query = """
            SELECT
                "순위",
                "브랜드명",
                "제품명",
                "가격",
                "날짜와시간",
                "채널",
                "카테고리",
                created_at
            FROM benow.yesstyle_table
            WHERE "채널" = %s
              AND "카테고리" = %s
              AND created_at = %s
            ORDER BY "순위" ASC
        """
        params: List = ["YesStyle", category, latest_created]
        if limit is not None:
            base_query += " LIMIT %s"
            params.append(limit)

        cursor.execute(base_query, params)
        rows = cursor.fetchall()
        if not rows:
            return None, latest_created, None

        df = pd.DataFrame(
            rows,
            columns=[
                COL_RANK,
                "브랜드명",
                "제품명",
                COL_PRICE,
                "날짜와시간",
                COL_CHANNEL,
                COL_CATEGORY,
                COL_COLLECTED_AT,
            ],
        )
        df.rename(
            columns={
                "브랜드명": COL_BRAND,
                "제품명": COL_PRODUCT_NAME,
                "날짜와시간": COL_DATETIME_TEXT,
            },
            inplace=True,
        )
        df[COL_RANK] = pd.to_numeric(df[COL_RANK], errors="coerce")
        df = df.dropna(subset=[COL_RANK])
        df[COL_RANK] = df[COL_RANK].astype(int)
        df.sort_values(by=COL_RANK, inplace=True)
        df.reset_index(drop=True, inplace=True)

        label = latest_created.strftime("%y년 %m월 %d일 %H시") if latest_created else None

        return df, latest_created, label
    except Error as exc:
        print(f"PostgreSQL 조회 오류: {exc}")
        return None, None, None
    finally:
        if connection:
            connection.close()


# --------------------------------------------------------------------------- #
# 데이터 가공
# --------------------------------------------------------------------------- #
def _format_status_from_delta(delta_value) -> str:
    if pd.isna(delta_value):
        return "신규"
    if delta_value > 0:
        return f"▲{int(delta_value)}"
    if delta_value < 0:
        return f"▼{abs(int(delta_value))}"
    return "-"


def _preprocess_rank_dataframe(df: Optional[pd.DataFrame]) -> Optional[pd.DataFrame]:
    if df is None:
        return None
    work = df.copy()
    if COL_PRODUCT_NAME in work.columns:
        work = work[work[COL_PRODUCT_NAME].notna()]
        work[COL_PRODUCT_NAME] = work[COL_PRODUCT_NAME].apply(_normalize_product_name)
    if COL_RANK in work.columns:
        work[COL_RANK] = pd.to_numeric(work[COL_RANK], errors="coerce")
        work = work[work[COL_RANK].notna()]
        work[COL_RANK] = work[COL_RANK].astype(int)
    for col in (COL_BRAND, COL_CHANNEL, COL_CATEGORY):
        if col in work.columns:
            work[col] = (
                work[col]
                .fillna("")
                .apply(lambda x: re.sub(r"\s+", " ", str(x)).strip())
            )
    _add_match_keys_inplace(work)
    return work


def annotate_rank_changes(
    current_df: pd.DataFrame,
    previous_df: Optional[pd.DataFrame],
) -> Tuple[pd.DataFrame, Optional[pd.DataFrame], Optional[str]]:
    current_clean = _preprocess_rank_dataframe(current_df)
    previous_clean = _preprocess_rank_dataframe(previous_df)

    if previous_clean is None or previous_clean.empty:
        annotated = current_clean.copy()
        annotated[COL_PREVIOUS_RANK] = pd.NA
        annotated[COL_RANK_DELTA] = pd.NA
        annotated[COL_STATUS] = "신규"
        annotated.sort_values(by=[COL_CATEGORY, COL_RANK], inplace=True)
        return annotated, None, None

    keys = [COL_CHANNEL, COL_CATEGORY, COL_BRAND_KEY, COL_PRODUCT_KEY]
    merged = current_clean.merge(
        previous_clean[keys + [COL_RANK]],
        how="left",
        on=keys,
        suffixes=("", "_previous"),
    )

    merged.rename(columns={f"{COL_RANK}_previous": COL_PREVIOUS_RANK}, inplace=True)
    merged[COL_PREVIOUS_RANK] = pd.to_numeric(merged[COL_PREVIOUS_RANK], errors="coerce").astype("Int64")
    merged[COL_RANK_DELTA] = merged[COL_PREVIOUS_RANK] - merged[COL_RANK]
    merged[COL_RANK_DELTA] = pd.to_numeric(merged[COL_RANK_DELTA], errors="coerce").astype("Int64")

    merged[COL_STATUS] = merged[COL_RANK_DELTA].apply(_format_status_from_delta)

    dropped = previous_clean.merge(
        merged[keys],
        how="left",
        on=keys,
        indicator=True,
    )
    dropped = dropped[dropped["_merge"] == "left_only"].drop(columns=["_merge"])
    if dropped.empty:
        dropped_df = None
    else:
        dropped_df = dropped.sort_values(by=[COL_CATEGORY, COL_RANK]).copy()
        dropped_df[COL_PREVIOUS_RANK] = dropped_df[COL_RANK]
        dropped_df.drop(columns=[COL_RANK], inplace=True)
        dropped_df[COL_STATUS] = "차트 아웃"

    previous_label = None
    if previous_df is not None and COL_DATETIME_TEXT in previous_df.columns:
        series = previous_df[COL_DATETIME_TEXT].dropna()
        if not series.empty:
            previous_label = series.iloc[0]

    merged.sort_values(by=[COL_CATEGORY, COL_RANK], inplace=True)
    return merged, dropped_df, previous_label


def add_weekly_rank_changes(
    current_df: pd.DataFrame,
    previous_week_df: Optional[pd.DataFrame],
) -> Tuple[pd.DataFrame, Optional[str]]:
    if previous_week_df is None or previous_week_df.empty:
        df = current_df.copy()
        df[COL_PREVIOUS_RANK_WEEK] = pd.NA
        df[COL_WEEK_DELTA] = pd.NA
        df[COL_WEEK_STATUS] = "-"
        return df, None

    prev_clean = _preprocess_rank_dataframe(previous_week_df)
    if prev_clean is None or prev_clean.empty:
        df = current_df.copy()
        df[COL_PREVIOUS_RANK_WEEK] = pd.NA
        df[COL_WEEK_DELTA] = pd.NA
        df[COL_WEEK_STATUS] = "-"
        return df, None

    keys = [COL_CHANNEL, COL_CATEGORY, COL_BRAND_KEY, COL_PRODUCT_KEY]
    merged = current_df.merge(
        prev_clean[keys + [COL_RANK]],
        how="left",
        on=keys,
        suffixes=("", "_week"),
    )

    merged.rename(columns={f"{COL_RANK}_week": COL_PREVIOUS_RANK_WEEK}, inplace=True)
    merged[COL_PREVIOUS_RANK_WEEK] = pd.to_numeric(merged[COL_PREVIOUS_RANK_WEEK], errors="coerce").astype("Int64")
    merged[COL_WEEK_DELTA] = merged[COL_PREVIOUS_RANK_WEEK] - merged[COL_RANK]
    merged[COL_WEEK_DELTA] = pd.to_numeric(merged[COL_WEEK_DELTA], errors="coerce").astype("Int64")
    merged[COL_WEEK_STATUS] = merged[COL_WEEK_DELTA].apply(_format_status_from_delta)

    week_label = None
    if previous_week_df is not None and COL_DATETIME_TEXT in previous_week_df.columns:
        series = previous_week_df[COL_DATETIME_TEXT].dropna()
        if not series.empty:
            week_label = series.iloc[0]

    merged.sort_values(by=[COL_CATEGORY, COL_RANK], inplace=True)
    return merged, week_label


def summarize_changes_for_email(
    annotated_df: pd.DataFrame,
    dropped_df: Optional[pd.DataFrame],
    run_time: datetime,
    previous_label: Optional[str],
    previous_week_label: Optional[str],
) -> str:
    lines: List[str] = [
        "안녕하세요 위마케팅 사업지원유닛 여도현입니다.",
        "금일자 YesStyle 전일·전주 대비 랭킹 변동 현황 보고드립니다.",
        "",
        f"수집 일시 (KST): {run_time.strftime('%Y-%m-%d %H:%M')}",
    ]
    if previous_label:
        lines.append(f"전일 기준: {previous_label}")
    else:
        lines.append("전일 기준: 참고 가능한 데이터 없음")
    if previous_week_label:
        lines.append(f"전주 기준: {previous_week_label}")
    else:
        lines.append("전주 기준: 참고 가능한 데이터 없음")
    lines.append("")

    for category in annotated_df[COL_CATEGORY].dropna().unique():
        cat_df = annotated_df[annotated_df[COL_CATEGORY] == category].copy()
        cat_df = cat_df[cat_df[COL_RANK] <= REPORT_LIMIT]
        cat_df["_delta"] = pd.to_numeric(cat_df[COL_RANK_DELTA], errors="coerce")

        up_count = int((cat_df["_delta"] > 0).sum())
        down_count = int((cat_df["_delta"] < 0).sum())
        new_count = int((cat_df[COL_STATUS] == "신규").sum())
        keep_count = int((cat_df[COL_STATUS] == "-").sum())

        lines.append(f"[{category}] ▲ {up_count} / ▼ {down_count} / 신규 {new_count} / 유지 {keep_count}")

        highlights: List[str] = []
        seen: set[Tuple[str, str]] = set()

        def _append_rows(df: pd.DataFrame):
            for _, r in df.iterrows():
                brand = r.get(COL_BRAND, "")
                product = r.get(COL_PRODUCT_NAME, "")
                key = (brand, product)
                status = str(r.get(COL_STATUS, "") or "")
                if key in seen or not status or status == "-":
                    continue
                highlights.append(f"  {status}: {brand} - {product}(#{int(r[COL_RANK])})")
                seen.add(key)
                if len(highlights) >= 7:
                    break

        cat_df["_abs_delta"] = cat_df["_delta"].abs()
        primary = cat_df[cat_df["_abs_delta"] >= 2].sort_values(by=["_abs_delta", COL_RANK], ascending=[False, True])
        _append_rows(primary)

        if len(highlights) < 5:
            brand_mask = cat_df[COL_BRAND].fillna("").str.lower()
            priority = cat_df[
                brand_mask.isin({"numbuzin", "fwee"})
                & (
                    cat_df["_abs_delta"].fillna(0) > 0
                    | (cat_df[COL_STATUS] == "신규")
                )
            ].sort_values(by=["_abs_delta", COL_RANK], ascending=[False, True])
            _append_rows(priority)

        if len(highlights) < 7:
            top10 = cat_df[(cat_df[COL_RANK] <= 10) & (cat_df[COL_STATUS] != "-")].sort_values(by=COL_RANK)
            _append_rows(top10)

        lines.extend(highlights)
        if highlights:
            lines.append("")

    if dropped_df is not None and not dropped_df.empty:
        filtered = dropped_df[
            dropped_df[COL_PREVIOUS_RANK].notna()
            & (dropped_df[COL_PREVIOUS_RANK] <= REPORT_LIMIT)
        ].copy()
        if not filtered.empty:
            lines.append(f"차트 아웃 {len(filtered)}개")
            top_drop = filtered.sort_values(COL_PREVIOUS_RANK).head(5)
            for _, row in top_drop.iterrows():
                previous_rank = row.get(COL_PREVIOUS_RANK)
                rank_text = f"#{int(previous_rank)}" if pd.notna(previous_rank) else "-"
                lines.append(f"  {rank_text} {row.get(COL_BRAND, '')} - {row.get(COL_PRODUCT_NAME, '')}")
    
    lines.append("")
    lines.append("감사합니다.")
    lines.append("여도현 드림")
    return "\n".join(lines)

def _format_korean_time(dt: datetime) -> str:
    hour = dt.hour
    minute = dt.minute

    if hour == 0:
        period = "오전"
        display_hour = 12
    elif 1 <= hour < 12:
        period = "오전"
        display_hour = hour
    elif hour == 12:
        period = "오후"
        display_hour = 12
    else:
        period = "오후"
        display_hour = hour - 12

    if minute:
        return f"{period} {display_hour}시 {minute:02d}분"
    return f"{period} {display_hour}시"


def save_excel_report(
    annotated_df: pd.DataFrame,
    raw_df: pd.DataFrame,
    previous_raw_df: Optional[pd.DataFrame],
    run_time: datetime,
    previous_label: Optional[str],
    previous_week_label: Optional[str],
    top_n: int = REPORT_LIMIT,
):
    report_df = (
        annotated_df.copy()
        .dropna(subset=[COL_RANK])
        .sort_values(by=[COL_CATEGORY, COL_RANK])
    )
    report_df = report_df.groupby(COL_CATEGORY, group_keys=False).head(top_n)

    display_df = report_df[
        [
            COL_RANK,
            COL_BRAND,
            COL_PRODUCT_NAME,
            COL_PRICE,
            COL_STATUS,
            COL_WEEK_STATUS,
        ]
    ].copy()
    display_df.rename(
        columns={
            COL_RANK: "랭킹",
            COL_BRAND: "브랜드",
            COL_PRODUCT_NAME: "제품명",
            COL_PRICE: "할인가",
            COL_STATUS: "전일대비 증감",
            COL_WEEK_STATUS: "전주대비 증감",
        },
        inplace=True,
    )
    display_df["전일대비 증감"] = display_df["전일대비 증감"].fillna("-")
    display_df["전주대비 증감"] = display_df["전주대비 증감"].fillna("-")
    display_df["비고"] = ""

    sheet_name = "Bestsellers"
    start_row = 5
    start_col = 3

    # 파일은 만들지 않고 메모리에서만 Excel만들기
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        display_df.to_excel(
            writer,
            index=False,
            sheet_name=sheet_name,
            startrow=start_row,
            startcol=start_col,
        )

        worksheet = writer.sheets[sheet_name]
        worksheet["D2"] = "■ 예스타일 랭킹 변동사항"
        worksheet["D2"].font = Font(bold=True)
        worksheet["D2"].alignment = Alignment(horizontal="left")

        worksheet["D4"] = f"- 기준일자 : {run_time.strftime('%m/%d')} ({_format_korean_time(run_time)})"
        worksheet["D4"].alignment = Alignment(horizontal="left")
        header_row = start_row + 1
        header_font = Font(bold=True)
        for cells in worksheet.iter_rows(
            min_row=header_row,
            max_row=header_row,
            min_col=start_col + 1,
            max_col=start_col + len(display_df.columns),
        ):
            for cell in cells:
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

        data_row_start = header_row + 1
        highlight_terms = ["numbuzin", "fwee", "knock"]
        highlight_terms = [re.sub(r"\s+", "", term).lower() for term in highlight_terms]
        highlight_fill = PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00")
        brand_col_letter = get_column_letter(start_col + 1 + 1)
        for row in worksheet.iter_rows(
            min_row=data_row_start,
            max_row=data_row_start + len(display_df) - 1,
            min_col=start_col + 1,
            max_col=start_col + len(display_df.columns),
        ):
            row_idx = row[0].row
            brand_cell = worksheet[f"{brand_col_letter}{row_idx}"]
            brand_value = str(brand_cell.value) if brand_cell.value is not None else ""
            brand_key = re.sub(r"\s+", "", brand_value).lower()
            if any(term in brand_key for term in highlight_terms):
                for cell in row:
                    cell.fill = highlight_fill
            for idx, cell in enumerate(row):
                if idx in (0, 4, 5):
                    cell.alignment = Alignment(horizontal="center")
                elif idx == 3:
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")

        for offset in (4, 5):
            trend_col = get_column_letter(start_col + 1 + offset)
            for row_idx in range(data_row_start, data_row_start + len(display_df)):
                cell = worksheet[f"{trend_col}{row_idx}"]
                value = str(cell.value) if cell.value is not None else ""
                if value.startswith("▲"):
                    cell.font = Font(color="FF0000")
                elif value.startswith("▼"):
                    cell.font = Font(color="0000FF")

        column_widths = [8, 22, 40, 18, 12, 12, 18]
        for offset, width in enumerate(column_widths):
            worksheet.column_dimensions[get_column_letter(start_col + 1 + offset)].width = width

        worksheet.freeze_panes = f"{get_column_letter(start_col + 1)}{header_row + 1}"
        output.seek(0)
    # 엑셀 내용물 받기
    return output.getvalue()

# --------------------------------------------------------------------------- #
# 메인 로직
# --------------------------------------------------------------------------- #
def main():
    options, temp_profile = build_chrome_options(profile_prefix="yesstyle_profile_")
    chrome_bin = os.getenv("CHROME_BIN")
    if chrome_bin:
        options.binary_location = chrome_bin

    page_load_timeout = int(os.getenv("SELENIUM_PAGELOAD_TIMEOUT", "60"))
    implicit_wait = int(os.getenv("SELENIUM_IMPLICIT_WAIT", "10"))
    driver = create_chrome_driver(
        options=options,
        temp_profile=temp_profile,
        page_load_timeout=page_load_timeout,
        implicit_wait=implicit_wait,
    )

    run_time = datetime.now(SEOUL_TZ)
    run_time_naive = run_time.astimezone(SEOUL_TZ).replace(tzinfo=None)
    date_str = f"{run_time.strftime('%y')}년 {run_time.month}월 {run_time.day}일 {run_time.hour}시"
    iso_timestamp = run_time.strftime("%Y-%m-%d %H:%M:%S")

    category_configs = [
        {"name": "Skincare", "url": SKINCARE_URL, "limit": CRAWL_LIMIT},
        {"name": "Numbuzin", "url": NUMBUZIN_URL, "limit": CRAWL_LIMIT},
        {"name": "Bestsellers", "url": BESTSELLER_URL, "limit": CRAWL_LIMIT},
    ]

    category_frames: dict[str, pd.DataFrame] = {}
    try:
        for config in category_configs:
            print(f"Start crawling {config['name']}...")
            names, prices, ranks, brands = yesstyle_scroll_crawling(
                driver,
                config["url"],
                target_count=config["limit"],
            )
            df = pd.DataFrame(
                {
                    COL_RANK: ranks,
                    COL_BRAND: brands,
                    COL_PRODUCT_NAME: names,
                    COL_PRICE: prices,
                    COL_DATETIME_TEXT: [date_str] * len(names),
                    COL_CHANNEL: ["YesStyle"] * len(names),
                    COL_CATEGORY: [config["name"]] * len(names),
                }
            )
            df[COL_COLLECTED_AT] = iso_timestamp
            _add_match_keys_inplace(df)
            category_frames[config["name"]] = df.dropna(subset=[COL_RANK, COL_PRODUCT_NAME]).copy()
    finally:
        cleanup_driver(driver)

    for df in category_frames.values():
        df[COL_RANK] = pd.to_numeric(df[COL_RANK], errors="coerce")
        df.dropna(subset=[COL_RANK, COL_PRODUCT_NAME], inplace=True)
        df[COL_RANK] = df[COL_RANK].astype(int)
        for col in (COL_BRAND, COL_PRICE, COL_CHANNEL, COL_CATEGORY):
            if col in df.columns:
                df[col] = df[col].fillna("")
        df.sort_values(by=COL_RANK, inplace=True)
        df.reset_index(drop=True, inplace=True)
        _add_match_keys_inplace(df)

    combined_df_all = pd.concat(category_frames.values(), ignore_index=True)

    data_dir = ensure_data_directory()
    prev_df, prev_created_at, previous_label = fetch_previous_snapshot_from_db(
        "Bestsellers", run_time_naive - timedelta(days=1)
    )
    if prev_created_at:
        print(f"Loaded previous snapshot: {prev_created_at}")
    else:
        print("No previous snapshot found.")

    prev_week_df, prev_week_created_at, previous_week_label = fetch_previous_snapshot_from_db(
        "Bestsellers", run_time_naive - timedelta(days=7)
    )
    if prev_week_created_at:
        print(f"Loaded previous week snapshot: {prev_week_created_at}")
    else:
        print("No previous week snapshot found.")

    if prev_df is not None:
        prev_df = _deduplicate_by_match_keys(prev_df, limit=CRAWL_LIMIT)
    if prev_week_df is not None:
        prev_week_df = _deduplicate_by_match_keys(prev_week_df, limit=CRAWL_LIMIT)

    bestsellers_df = category_frames["Bestsellers"]
    dedup_order = bestsellers_df.sort_values([COL_BRAND_KEY, COL_PRODUCT_KEY, COL_RANK])
    dedup_order = dedup_order.drop_duplicates(subset=[COL_BRAND_KEY, COL_PRODUCT_KEY], keep="last")
    bestsellers_df = (
        dedup_order.sort_values(by=COL_RANK)
        .head(CRAWL_LIMIT)
        .reset_index(drop=True)
    )
    category_frames["Bestsellers"] = bestsellers_df.copy()
    annotated_df, dropped_df, previous_label = annotate_rank_changes(bestsellers_df, prev_df)
    annotated_df, previous_week_label = add_weekly_rank_changes(annotated_df, prev_week_df)

    created_at_value = run_time_naive
    db_rows = [
        (
            int(row[COL_RANK]),
            row[COL_BRAND],
            row[COL_PRODUCT_NAME],
            row[COL_PRICE],
            row[COL_DATETIME_TEXT],
            row[COL_CHANNEL],
            row[COL_CATEGORY],
            None,
            None,
            None,
            None,
            created_at_value,
        )
        for _, row in combined_df_all.iterrows()
    ]

    timestamp = run_time.strftime("%y%m%d_%H%M")
    result_excel = save_excel_report(
        annotated_df,
        bestsellers_df,
        prev_df,
        run_time=run_time_naive,
        previous_label=previous_label,
        previous_week_label=previous_week_label,
        top_n=REPORT_LIMIT,
    )

    client = WebClient(token=SLACK_BOT_TOKEN)
    resp = client.files_upload_v2(
        channel="C0ABHFQKY8L",
        filename="Yesstyle.xlsx",
        content = result_excel,
    )

    # def upload_excel_to_slack():
    #     response = requests.post(
    #         "https://slack.com/api/files.upload",
    #         headers={
    #             "Authorization":f"Bearer {SLACK_BOT_TOKEN}"
    #         },
    #         files={
    #                 "file": excel
    #         },
    #         data={
    #         "channels": "#프로젝트_테스트_md",
    #         "title": "YesStyle 크롤링 결과",
    #         }
    #     )       
    #     return response.json()
    
    return resp
    print("YesStyle 크롤링 완료.")
    

